from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from core.models import OctoTaskRule


SEVERITY_MAP = {
    "Kritik": "critical",
    "Uyari": "warning",
    "Uyarı": "warning",
    "Bilgi": "info",
    "Fırsat": "opportunity",
    "Firsat": "opportunity",
}

MODULE_MAP = {
    "Campaign": "performance",
    "AdGroup": "performance",
    "Ad": "performance",
    "Creative": "creative",
    "Competitor": "competitor",
    "Budget": "budget",
    "Conversion": "conversion",
    "Audience": "performance",
    "Platform": "performance",
}


class Command(BaseCommand):
    help = "Master Matrix Excel dosyasından OctoTaskRule seed eder."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        file_path = options["file"]
        dry_run = options["dry_run"]

        self.stdout.write("Excel açılıyor...")

        wb = load_workbook(
            file_path,
            read_only=True,
            data_only=True
        )

        sheet_name = "Master Matrix"

        if sheet_name not in wb.sheetnames:
            raise CommandError("Master Matrix sayfası bulunamadı.")

        ws = wb[sheet_name]

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        headers = {}
        for index, value in enumerate(header_row):
            if value:
                headers[str(value).strip()] = index

        required_columns = [
            "Görev ID",
            "Bildirim Seviyesi",
            "Modül",
            "Platform",
            "Ana Tablo",
            "Teknik Koşul",
            "Kullanıcı Diliyle Koşul",
            "Kök Neden ve İncelenecek Noktalar",
            "Nokta Atışı Aksiyon",
            "Beklenen Sonuç",
            "Öncelik Puanı",
            "Görev Kartı Başlığı",
            "Görev Kartı Açıklaması",
            "Buton / CTA",
        ]

        missing = [col for col in required_columns if col not in headers]

        if missing:
            raise CommandError("Eksik kolon var: " + ", ".join(missing))

        created_count = 0
        updated_count = 0
        skipped_count = 0
        read_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            code = self.get_value(row, headers, "Görev ID")

            if not code:
                skipped_count += 1
                continue

            raw_severity = self.get_value(row, headers, "Bildirim Seviyesi")
            raw_module = self.get_value(row, headers, "Modül")

            severity = SEVERITY_MAP.get(raw_severity, "info")
            module = MODULE_MAP.get(raw_module, "performance")

            title_tr = self.get_value(row, headers, "Görev Kartı Başlığı")
            message_tr = self.get_value(row, headers, "Görev Kartı Açıklaması")
            action_text_tr = self.get_value(row, headers, "Nokta Atışı Aksiyon")

            user_condition = self.get_value(row, headers, "Kullanıcı Diliyle Koşul")
            root_cause = self.get_value(row, headers, "Kök Neden ve İncelenecek Noktalar")
            expected_result = self.get_value(row, headers, "Beklenen Sonuç")
            condition_description = self.get_value(row, headers, "Teknik Koşul")

            cta_text = self.get_value(row, headers, "Buton / CTA")
            source_platform = self.get_value(row, headers, "Platform")
            source_table = self.get_value(row, headers, "Ana Tablo")
            priority_score = self.get_int_value(row, headers, "Öncelik Puanı")

            if not title_tr:
                title_tr = action_text_tr or f"Octo görevi {code}"

            if not message_tr:
                message_tr = user_condition or root_cause or title_tr

            defaults = {
                "module": module,
                "severity": severity,
                "title_tr": title_tr[:255],
                "message_tr": message_tr,
                "action_text_tr": action_text_tr[:255] if action_text_tr else None,
                "title_en": None,
                "message_en": None,
                "action_text_en": None,
                "condition_key": str(code).lower().replace("-", "_"),
                "condition_description": condition_description or None,
                "root_cause": root_cause or None,
                "expected_result": expected_result or None,
                "cta_text": cta_text[:255] if cta_text else None,
                "user_condition": user_condition or None,
                "source_platform": source_platform[:100] if source_platform else None,
                "source_table": source_table[:100] if source_table else None,
                "priority_score": priority_score,
                "is_active": True,
            }

            read_count += 1

            if dry_run:
                if OctoTaskRule.objects.filter(code=code).exists():
                    updated_count += 1
                else:
                    created_count += 1
            else:
                _, created = OctoTaskRule.objects.update_or_create(
                    code=code,
                    defaults=defaults,
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            if read_count % 250 == 0:
                self.stdout.write(f"{read_count} satır işlendi...")

        wb.close()

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: Veritabanına yazılmadı."))

        self.stdout.write(self.style.SUCCESS("Octo görev seed işlemi tamamlandı."))
        self.stdout.write(f"Okunan: {read_count}")
        self.stdout.write(f"Oluşturulacak/Oluşturulan: {created_count}")
        self.stdout.write(f"Güncellenecek/Güncellenen: {updated_count}")
        self.stdout.write(f"Atlanan: {skipped_count}")

    def get_value(self, row, headers, column_name):
        index = headers[column_name]

        if index >= len(row):
            return ""

        value = row[index]

        if value is None:
            return ""

        return str(value).strip()

    def get_int_value(self, row, headers, column_name):
        value = self.get_value(row, headers, column_name)

        if not value:
            return 50

        try:
            return int(float(value))
        except Exception:
            return 50